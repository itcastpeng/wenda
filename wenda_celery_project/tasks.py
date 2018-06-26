#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author:zhangcong
# Email:zc_92@sina.com

from __future__ import absolute_import, unicode_literals
import hashlib
from .celery import app
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.exceptions import IllegalCharacterError
import sys
import os
from webadmin.modules import SendMsg
from webadmin.modules import WeChat

# project_dir 是
project_dir = os.path.dirname(os.getcwd())
sys.path.append(project_dir)
os.environ['DJANGO_SETTINGS_MODULE'] = 'wenda.settings'
import django

django.setup()
import requests
import xlrd,time
from django.db.utils import IntegrityError
from webadmin.modules import RedisOper
import sys ,datetime ,os ,django, json
from django.db.models import Q, Count
project_dir = os.path.dirname(os.getcwd())
sys.path.append(project_dir)
print(project_dir)
os.environ['DJANGO_SETTINGS_MODULE'] = 'wenda.settings'
django.setup()
from webadmin import models
from webadmin.modules.WeChat import WeChatPublicSendMsg
from time import sleep

# 客户首次创建任务的时候,将客户提交的 excel 表格的数据取出来然后写入到新的 excel 表格中, 在第一列新增 问答地址链接
@app.task
def CreateExcel(excel_data, file_save_path):
    wb = Workbook()

    ws = wb.active
    ws.cell(row=1, column=1, value="新问答")
    ws.cell(row=2, column=1, value="问答地址链接")
    ws.cell(row=2, column=2, value="问题")
    ws.cell(row=2, column=3, value="答案")

    ft1 = Font(name='宋体', size=28)
    a1 = ws['A1']
    a1.font = ft1

    # 合并单元格   a1 - c1
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=3)

    # 设置列宽
    ws.column_dimensions['A'].width = 63
    ws.column_dimensions['B'].width = 37
    ws.column_dimensions['C'].width = 39

    # 设置行高
    ws.row_dimensions[2].height = 50

    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')

    # [['', '问题1', '答案1'], ['', '问题2', '答案2']]

    for row_line, row in enumerate(excel_data, start=3):
        # row_line 第几行
        # row  数据
        for col_index, col in enumerate(row, start=1):
            # col_index 第几列
            # col  数据

            print(row_line, col_index, col)

            ws.cell(row=row_line, column=col_index, value=col)

    wb.save(file_save_path)


# 统计是否有编辑或机器人提交发布问答的表格
@app.task
def CheckWenda():
    task_objs = models.Task.objects.select_related('publish_user').filter(
        is_delete=False,
        is_check=False,
        status=7,
        release_user__is_delete=False,
        publish_user__isnull=False,
        publish_task_result_file_path__isnull=False,
    )
    print(task_objs.count())
    for obj in task_objs:
        # 先清空该任务在回链表中的所有链接
        models.WendaLink.objects.filter(task=obj).delete()

        # 将提交的 excel 表格中的数据读出
        print(obj.name)
        book = xlrd.open_workbook(obj.publish_task_result_file_path)
        sh = book.sheet_by_index(0)

        query = []

        for row in range(2, sh.nrows):
            url = sh.cell_value(rowx=row, colx=0)
            keywords = sh.cell_value(rowx=row, colx=1)

            wenda_link_objs = models.WendaLink.objects.filter(url=url)
            if wenda_link_objs.count() > 0:  # 如果该链接已经存在,则跳过
                continue

            query.append(
                models.WendaLink(
                    task=obj,
                    url=url,
                    check_date=datetime.datetime.now() + datetime.timedelta(days=3),
                )
            )
            try:
                models.SearchKeywordsRank.objects.create(
                    client_user_id=obj.release_user.id,
                    task=obj,
                    keywords=keywords,
                    type=2,
                    oper_user_id=8
                )
            except IntegrityError:
                continue

        models.WendaLink.objects.bulk_create(query)
        obj.is_check = True
        obj.save()


# 检测任务是否存在异常
@app.task
def CheckTaskYichang():
    task_objs = models.Task.objects.filter(is_delete=False, status__gte=6, release_user__is_delete=False)
    print("task_objs --> ", len(task_objs))
    for obj in task_objs:
        if obj.wendalink_set.filter(status__gt=2):

            if obj.is_yichang:  # 如果已经是异常状态,则跳过
                continue

            yichang = True
        else:
            yichang = False

        obj.is_yichang = yichang
        obj.yichang_date = datetime.datetime.now()
        obj.save()


# 将数据写入到 excel 中, 问答机器人中生成报表
@app.task
def WendaRobotWriteExcel(file_name, wenda_robot_task_list, status_choices):
    print(file_name)
    wb = Workbook()
    ws = wb.active

    for index, obj in enumerate(wenda_robot_task_list, start=1):
        status = ''
        for i in status_choices:
            if obj["status"] == i[0]:
                status = i[1]
                break

        ws.cell(row=index, column=1, value=status)
        ws.cell(row=index, column=2, value=obj["wenda_url"])
        ws.cell(row=index, column=3, value=obj["title"])
        ws.cell(row=index, column=4, value=obj["content"])

    wb.save(file_name)


# 编辑内容详情中 商务通角色生成报表
@app.task
def EditPublickTaskManagementWriteExcel(file_name, data_list):
    print(file_name)
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="问题")
    ws.cell(row=1, column=2, value="答案")
    ws.cell(row=1, column=3, value="知道地址")
    ws.cell(row=1, column=4, value="状态")
    ws.cell(row=1, column=5, value="创建时间")
    ws.cell(row=1, column=6, value="最后查询时间")

    for index, obj in enumerate(data_list, start=1):
        ws.cell(row=index, column=1, value=obj["title"])
        ws.cell(row=index, column=2, value=obj["content"])
        ws.cell(row=index, column=3, value=obj["url"])
        ws.cell(row=index, column=4, value=obj["status"])
        ws.cell(row=index, column=5, value=obj["create_date"])
        ws.cell(row=index, column=6, value=obj["update_date"])

    wb.save(file_name)


# 将分配给机器人的任务加入到机器人任务表中
@app.task
def ToRobotTask():
    robot_ids = [i[0] for i in models.UserProfile.objects.filter(role_id=10).values_list('id')]

    task_objs = models.Task.objects.filter(status=6, publish_user_id__in=robot_ids)  # 状态为发布中,且发布用户为机器人
    for obj in task_objs:
        wenda_robot_task_obj = models.WendaRobotTask.objects.filter(task=obj)
        if wenda_robot_task_obj:  # 已经进入机器人中则跳过
            continue

        obj.fabu_date = datetime.datetime.now()
        obj.save()

        file_path = obj.task_result_file_path
        if not file_path:
            file_path = obj.task_demand_file_path

        # 从excel 表格中读取数据
        filename = os.path.join(project_dir, "wenda", file_path)
        print(filename)
        book = xlrd.open_workbook(filename=filename)
        sh = book.sheet_by_index(0)

        for row in range(2, sh.nrows):
            url = sh.cell_value(rowx=row, colx=0)
            title = sh.cell_value(rowx=row, colx=1)
            content = sh.cell_value(rowx=row, colx=2)

            status = 1
            if obj.wenda_type == 2:
                status = 2

            edit_task_management_objs = models.EditTaskManagement.objects.filter(
                task__task=obj,
                edit_user__role_id=14,  # 编辑角色为 外部编辑-商务通渠道
            )
            if edit_task_management_objs:  # 如果编辑角色为外部编辑-商务通渠道,则任务状态直接改为已完成
                status = 6

            wenda_robot_task_obj = models.WendaRobotTask.objects.create(
                task=obj,
                wenda_url=url,
                title=title,
                content=content,
                release_platform=obj.release_platform,
                wenda_type=obj.wenda_type,
                next_date=datetime.datetime.now() + datetime.timedelta(minutes=30),
                status=status,
                add_map=obj.add_map
            )
            try:
                t_id = sh.cell_value(rowx=row, colx=3)
                edit_publick_task_management_obj = models.EditPublickTaskManagement.objects.get(id=t_id)
                edit_publick_task_management_obj.run_task = wenda_robot_task_obj
                edit_publick_task_management_obj.save()
            except IndexError:
                pass


# 将机器人中已经操作完的任务提交到后台
@app.task
def RobotTaskToTask():
    robot_ids = [i[0] for i in models.UserProfile.objects.filter(role_id=10).values_list('id')]

    # 状态为发布中,且发布用户为机器人

    five_days_ago = datetime.datetime.now() - datetime.timedelta(days=5)
    # q = Q(Q(status=6) or Q(create_date__lt=five_days_ago))
    task_objs = models.Task.objects.filter(publish_user_id__in=robot_ids, status=6)
    for task_obj in task_objs:
        print(task_obj.name)

        # 判断所有子任务是否都完成
        wenda_robot_task_objs = models.WendaRobotTask.objects.filter(task=task_obj)

        # 如果分配给机器人的所有任务都已经完成或者创建时间距离现在5天
        if (wenda_robot_task_objs.count() and wenda_robot_task_objs.filter(
                status=6).count() == wenda_robot_task_objs.count()) or task_obj.fabu_date < five_days_ago:
            if task_obj.create_date < five_days_ago:
                wenda_robot_task_objs.update(status=6)
                models.EditPublickTaskManagement.objects.filter(task__task__task=task_obj).update(status=3)

            wb = Workbook()

            ws = wb.active
            ws.cell(row=1, column=1, value="新问答")
            ws.cell(row=2, column=1, value="问答地址链接")
            ws.cell(row=2, column=2, value="问题")
            ws.cell(row=2, column=3, value="答案")

            ft1 = Font(name='宋体', size=28)
            a1 = ws['A1']
            a1.font = ft1

            # 合并单元格   a1 - c1
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=3)

            # 设置列宽
            ws.column_dimensions['A'].width = 63
            ws.column_dimensions['B'].width = 37
            ws.column_dimensions['C'].width = 39

            # 设置行高
            ws.row_dimensions[2].height = 50

            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C2'].alignment = Alignment(horizontal='center', vertical='center')

            excel_data = []
            for wenda_robot_task_obj in wenda_robot_task_objs:
                excel_data.append(
                    [wenda_robot_task_obj.wenda_url, wenda_robot_task_obj.title, wenda_robot_task_obj.content])

            # [['', '问题1', '答案1'], ['', '问题2', '答案2']]

            for row_line, row in enumerate(excel_data, start=3):
                # row_line 第几行
                # row  数据
                for col_index, col in enumerate(row, start=1):
                    # col_index 第几列
                    # col  数据

                    print(row_line, col_index, col)

                    ws.cell(row=row_line, column=col_index, value=col)

            # 相对路径
            file_name = '/'.join(["statics", "task_excel", "publish", task_obj.name + ".xlsx"])

            # 绝对路径
            file_save_path = os.path.join(project_dir, "wenda", file_name)
            print(file_save_path)
            wb.save(file_save_path)

            task_obj.status = 7
            task_obj.publish_task_result_file_path = file_name
            task_obj.is_check = False
            task_obj.yichang_date = None

            message = "您的任务 [{task_name}] 编辑已经提交结果,请前去查看,并审核".format(
                task_name=task_obj.name,
            )

            task_process_log_objs = models.TaskProcessLog.objects.filter(
                task=task_obj,
                status=6,
                oper_user_id=8
            )
            # 如果没有创建过记录,则添加记录
            if task_process_log_objs.count() == 0:
                # 发布通知消息
                models.Message.objects.create(
                    user_id=task_obj.release_user.id,
                    content=message,
                    m_user_id=8,  # 机器人小明的账号是8
                )

                # 编辑提交任务结果
                models.TaskProcessLog.objects.create(
                    task=task_obj,
                    status=6,
                    oper_user_id=8
                )

            task_obj.update_date = datetime.datetime.now()
            task_obj.save()


# 发送微信企业号通知
@app.task
def send_msg(user_id, msg, w_type='text', url=None):
    send_msg_obj = SendMsg.SendMsg()
    if w_type == "text":
        send_msg_obj.send_text_msg(user_id, msg)
    elif w_type == "card":
        send_msg_obj.send_card_msg(user_id, msg, url)


# 发送微信公众号通知
@app.task
def send_msg_gongzhonghao(post_data):
    send_msg_obj = WeChat.WeChatPublicSendMsg()
    send_msg_obj.sendTempMsg(post_data)


# 编辑内容管理中, 营销顾问要将任务进入发布队列,需要创建excel表格
@app.task
def edit_content_management_create_excel(o_id, file_save_path, wenda_type):
    wb = Workbook()

    ws = wb.active
    if wenda_type in [1, 10]:
        ws.cell(row=1, column=1, value="新问答")
    else:
        ws.cell(row=1, column=1, value="老问答")
    ws.cell(row=2, column=1, value="问答地址链接")
    ws.cell(row=2, column=2, value="问题")
    ws.cell(row=2, column=3, value="答案")
    ws.cell(row=2, column=4, value="任务编号")

    ft1 = Font(name='宋体', size=28)
    a1 = ws['A1']
    a1.font = ft1

    # 合并单元格   a1 - c1
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=3)

    # 设置列宽
    ws.column_dimensions['A'].width = 63
    ws.column_dimensions['B'].width = 37
    ws.column_dimensions['C'].width = 39

    # 设置行高
    ws.row_dimensions[2].height = 50

    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')

    row_line = 3
    edit_publick_task_management_objs = models.EditPublickTaskManagement.objects.select_related(
        'task__task__client_user').filter(task__task_id=o_id)
    for obj in edit_publick_task_management_objs:
        try:
            if wenda_type == 2:
                ws.cell(row=row_line, column=1, value=obj.url)
            ws.cell(row=row_line, column=2, value=obj.title)
            ws.cell(row=row_line, column=3, value=obj.content)
            ws.cell(row=row_line, column=4, value=obj.id)
        except IllegalCharacterError:
            continue

        row_line += 1
    print(file_save_path)
    wb.save(file_save_path)


# 生成异常报表
@app.task
def generate_error_excel(o_id, file_save_path):
    obj = models.Task.objects.get(id=o_id)
    file_name_path = os.path.join(os.getcwd(), obj.publish_task_result_file_path)
    book = xlrd.open_workbook(file_name_path)
    sh = book.sheet_by_index(0)
    table_data = {}
    for row in range(2, sh.nrows):
        line_data = []
        for col in range(sh.ncols):
            value = sh.cell_value(rowx=row, colx=col)

            ctype = sh.cell(row, col).ctype  # 表格的数据类型
            if ctype == 2 and int(str(value).split(".")[1]) == 0:
                value = int(value)
            line_data.append(value)

        url = sh.cell_value(rowx=row, colx=0)
        table_data[url] = line_data

    wendalink_objs = obj.wendalink_set.filter(status__gt=2)

    wb = Workbook()

    ws = wb.active
    ws.cell(row=1, column=1, value="编号")
    ws.cell(row=1, column=2, value="状态")
    ws.cell(row=1, column=3, value="问题")
    ws.cell(row=1, column=4, value="答案")
    ws.cell(row=1, column=5, value="链接")
    ws.cell(row=1, column=6, value="检查时间")

    for index, obj in enumerate(wendalink_objs, start=2):
        ws.cell(row=index, column=1, value=index)
        ws.cell(row=index, column=2, value=obj.get_status_display())
        ws.cell(row=index, column=3, value=table_data[obj.url][1])
        ws.cell(row=index, column=4, value=table_data[obj.url][2])
        ws.cell(row=index, column=5, value=obj.url)
        ws.cell(row=index, column=6, value=obj.update_date.strftime("%y-%m-%d %H:%M:%S"))

    wb.save(file_save_path)


# 更新客户当日覆盖信息
@app.task
def update_client_day_cover():
    userprofile_objs = models.UserProfile.objects.filter(role_id=5, is_delete=False)

    data = {}

    print('userprofile_objs -->', userprofile_objs)
    for user_obj in userprofile_objs:
        data[user_obj.id] = {}
        print(user_obj.username)

        now_date = datetime.datetime.now().strftime("%Y-%m-%d")
        objs = models.SearchKeywordsRankLog.objects.select_related("keywords__client_user", "keywords").filter(
            keywords__client_user=user_obj,
            create_date=now_date
        )
        print('objs -->', objs)
        for obj in objs:
            create_date_str = obj.create_date.strftime("%Y-%m-%d")
            rank_num = len(obj.rank.split(","))

            if obj.task_type == 1:  # pc覆盖
                if create_date_str in data[user_obj.id]:
                    if obj.keywords.type == 1:  # 指定关键词
                        data[user_obj.id][create_date_str]["pc_cover"]["zhiding"] += rank_num
                    else:  # 问题关键词
                        data[user_obj.id][create_date_str]["pc_cover"]["wenti"] += rank_num
                else:
                    if obj.keywords.type == 1:  # 指定关键词
                        data[user_obj.id][create_date_str] = {
                            "pc_cover": {
                                "zhiding": rank_num,
                                "wenti": 0
                            },
                            "wap_cover": {
                                "zhiding": 0,
                                "wenti": 0
                            }
                        }
                    else:  # 问题关键词
                        data[user_obj.id][create_date_str] = {
                            "pc_cover": {
                                "zhiding": 0,
                                "wenti": rank_num
                            },
                            "wap_cover": {
                                "zhiding": 0,
                                "wenti": 0
                            }
                        }

            else:  # 移动覆盖
                if create_date_str in data[user_obj.id]:
                    if obj.keywords.type == 1:  # 指定关键词
                        data[user_obj.id][create_date_str]["wap_cover"]["zhiding"] += rank_num
                    else:  # 问题关键词
                        data[user_obj.id][create_date_str]["wap_cover"]["wenti"] += rank_num
                else:
                    if obj.keywords.type == 1:  # 指定关键词
                        data[user_obj.id][create_date_str] = {
                            "pc_cover": {
                                "zhiding": 0,
                                "wenti": 0
                            },
                            "wap_cover": {
                                "zhiding": rank_num,
                                "wenti": 0
                            }
                        }
                    else:  # 问题关键词
                        data[user_obj.id][create_date_str] = {
                            "pc_cover": {
                                "zhiding": 0,
                                "wenti": 0
                            },
                            "wap_cover": {
                                "zhiding": 0,
                                "wenti": rank_num
                            }
                        }
    print('data -->', data)
    for k, v in data.items():
        for k1, v1 in v.items():
            print(k, k1, v1)

            wap_cover_zhiding = v1["wap_cover"]["zhiding"]
            wap_cover_wenti = v1["wap_cover"]["wenti"]
            pc_cover_zhiding = v1["pc_cover"]["zhiding"]
            pc_cover_wenti = v1["pc_cover"]["wenti"]
            covering_total_number = wap_cover_zhiding + wap_cover_wenti + pc_cover_zhiding + pc_cover_wenti

            obj = models.ClientCoveringNumber.objects.filter(
                client_id=k,
                date=k1
            )

            if obj:
                obj.update(
                    covering_number=covering_total_number,
                    covering_zhiding_number_pc=pc_cover_zhiding,
                    covering_wenti_number_pc=pc_cover_wenti,
                    covering_zhiding_number_wap=wap_cover_zhiding,
                    covering_wenti_number_wap=wap_cover_wenti,
                )
            else:
                models.ClientCoveringNumber.objects.create(
                    client_id=k,
                    date=k1,
                    covering_number=covering_total_number,
                    covering_zhiding_number_pc=pc_cover_zhiding,
                    covering_wenti_number_pc=pc_cover_wenti,
                    covering_zhiding_number_wap=wap_cover_zhiding,
                    covering_wenti_number_wap=wap_cover_wenti,
                )


# 排名数据功能中生成下载报表
@app.task
def rank_data_generate_excel(file_name, data_list):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="医院名称")
    ws.cell(row=1, column=2, value="查询时间")
    ws.cell(row=1, column=3, value="类型")
    ws.cell(row=1, column=4, value="关键词")
    ws.cell(row=1, column=5, value="排名")
    ws.cell(row=1, column=6, value="关键词类型")

    for row, i in enumerate(data_list, start=2):
        ws.cell(row=row, column=1, value=i["username"])
        ws.cell(row=row, column=2, value=i["create_date"])
        ws.cell(row=row, column=3, value=i["task_type"])
        ws.cell(row=row, column=4, value=i["keywords"])
        ws.cell(row=row, column=5, value=i["rank"])
        ws.cell(row=row, column=6, value=i["type"])

        row += 1

    wb.save(file_name)


# 生成指定首页关键词覆盖报表 和 营销顾问能够直接上传的报表
@app.task
def keywords_top_page_cover_excel(user_id=None):
    if user_id:
        user_profile_objs = models.UserProfile.objects.filter(id=user_id)
    else:
        user_profile_objs = models.UserProfile.objects.filter(role_id=5, is_delete=False)

    for user_obj in user_profile_objs:

        print(user_obj.username)
        keywords_top_set_objs = models.KeywordsTopSet.objects.filter(client_user=user_obj)

        # 没有指定首页关键词的跳过
        if keywords_top_set_objs.count() == 0:
            continue
        else:
            wb = Workbook()
            ws = wb.active

            ws.cell(row=1, column=1, value="编号")
            ws.cell(row=1, column=2, value="客户名称")
            ws.cell(row=1, column=3, value="搜索关键词")
            ws.cell(row=1, column=4, value="知道标题")
            ws.cell(row=1, column=5, value="知道链接")
            ws.cell(row=1, column=6, value="排名位置")
            ws.cell(row=1, column=7, value="排名类型")
            ws.cell(row=1, column=8, value="是否采纳")
            ws.cell(row=1, column=9, value="创建时间")

            row = 2
            for keywords_top_set_obj in keywords_top_set_objs:
                for obj in keywords_top_set_obj.keywordstopinfo_set.all():

                    if obj.is_caina:
                        is_caina = "已采纳"
                    else:
                        is_caina = "未采纳"

                    try:
                        ws.cell(row=row, column=1, value=row - 1)
                        ws.cell(row=row, column=2, value=user_obj.username)
                        ws.cell(row=row, column=3, value=keywords_top_set_obj.keyword)
                        ws.cell(row=row, column=4, value=obj.title)
                        ws.cell(row=row, column=5, value=obj.url)
                        ws.cell(row=row, column=6, value=obj.rank)
                        ws.cell(row=row, column=7, value=obj.get_page_type_display())
                        ws.cell(row=row, column=8, value=is_caina)
                        ws.cell(row=row, column=9, value=keywords_top_set_obj.create_date.strftime("%y-%m-%d %H:%M:%S"))

                        row += 1
                    except IllegalCharacterError:
                        pass

            now_date = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
            keywords_top_page_cover_excel_path = "statics/task_excel/keywords_top_set/{name}_{now_date}.xlsx".format(
                name=user_obj.username,
                now_date=now_date
            )
            wb.save(os.path.abspath(keywords_top_page_cover_excel_path))
            user_obj.keywords_top_page_cover_excel_path = keywords_top_page_cover_excel_path

            # 生成营销顾问直接能够上传的报表
            keywords_top_info_objs = models.KeywordsTopInfo.objects.filter(
                keyword__client_user=user_obj).values("url", "title").annotate(Count("url")).order_by("url__count")
            guanjianci_leixing = ''

            if not keywords_top_info_objs:
                continue

            wb = Workbook()

            ws = wb.active
            ws.cell(row=1, column=1, value="编号")
            ws.cell(row=1, column=2, value="链接")
            ws.cell(row=1, column=3, value="问题")
            ws.cell(row=1, column=4, value="覆盖数")
            ws.cell(row=1, column=5, value="关键词")
            ws.cell(row=1, column=6, value="是否可以采纳")
            ws.cell(row=1, column=7, value="回复数量")
            ws.cell(row=1, column=8, value="创建时间")
            ws.cell(row=1, column=9, value="更新时间")
            ws.cell(row=1, column=10, value="状态")

            url_publish_list = [i[0] for i in models.WendaRobotTask.objects.filter(task__release_user=user_obj,
                wenda_type=2, ).values_list(
                "wenda_url")]
            n = 2
            for obj in keywords_top_info_objs:

                url = obj["url"]
                title = obj["title"]
                url__count = obj["url__count"]
                k_objs = models.KeywordsTopInfo.objects.select_related('keyword').filter(
                    keyword__client_user=user_obj,
                    url=url,
                ).order_by('update_date')

                if k_objs:
                    keywords_list = []
                    is_caina = ""
                    huifu_num = 0
                    create_date = ""
                    update_date = ''
                    for k_obj in k_objs:
                        guanjianci_leixing = k_obj.keyword.get_keywords_type_display()
                        keywords_list.append(k_obj.keyword.keyword + '__({guanjianci_leixing})'.format(
                            guanjianci_leixing=guanjianci_leixing))
                        keywords_list.append(k_obj.keyword.keyword)
                        is_caina = k_obj.is_caina
                        huifu_num = k_obj.huifu_num
                        create_date = k_obj.create_date.strftime("%Y-%m-%d")
                        if k_obj.update_date:
                            update_date = k_obj.update_date.strftime("%Y-%m-%d")

                    if is_caina:
                        is_caina = "采纳"
                    else:
                        is_caina = "未采纳"

                    if url in url_publish_list:
                        flag_publish_str = "已发布"
                    else:
                        flag_publish_str = "未发布"

                    keywords = "\n".join(keywords_list)

                    # print("-->", n, obj["id"], url, title, url__count, keywords, is_caina, huifu_num, flag_publish_str)
                    try:
                        ws.cell(row=n, column=1, value=n)
                        ws.cell(row=n, column=2, value=url)
                        ws.cell(row=n, column=3, value=title)
                        ws.cell(row=n, column=4, value=url__count)
                        ws.cell(row=n, column=5, value=keywords)
                        ws.cell(row=n, column=6, value=is_caina)
                        ws.cell(row=n, column=7, value=huifu_num)
                        ws.cell(row=n, column=8, value=create_date)
                        ws.cell(row=n, column=9, value=update_date)
                        ws.cell(row=n, column=10, value=flag_publish_str)

                        n += 1
                    except IllegalCharacterError:
                        pass

            keywords_top_page_cover_yingxiao_excel_path = "statics/task_excel/keywords_top_set/{name}_{now_date}.xlsx".format(
                name=user_obj.username,
                now_date=now_date
            )
            wb.save(os.path.abspath(keywords_top_page_cover_yingxiao_excel_path))
            user_obj.keywords_top_page_cover_yingxiao_excel_path = keywords_top_page_cover_yingxiao_excel_path

            models.UserProfile.objects.filter(id=user_obj.id).update(
                keywords_top_page_cover_yingxiao_excel_path=keywords_top_page_cover_yingxiao_excel_path,
                keywords_top_page_cover_excel_path=keywords_top_page_cover_excel_path
            )


# 统计3天未操作的客户,然后将客户名发送给对应的营销顾问    每天早上9点发送一次
@app.task
def tongji_kehu_shiyong():
    user_profile_objs = models.UserProfile.objects.filter(role_id=5, is_delete=False)
    up_time = datetime.datetime.now() - datetime.timedelta(days=3)

    client_username = ""
    for user_profile_obj in user_profile_objs:
        task_count = models.Task.objects.filter(release_user=user_profile_obj, update_date__gt=up_time).count()

        if task_count == 0:
            client_username += user_profile_obj.username + "\n"

    RedisOper.write_to_cache("tongji_kehu_shiyong", client_username)
    post_data = {
        "touser": "o7Xw_0bdwjqmqSsXBVGfZiYMy0pQ",  # 发送给张聪
        # "touser": "o7Xw_0ZgxSLdvmtFdOd4c9W4OAn4",   # 发送给寻梦
        # "touser": "o7Xw_0RM0Cw5IU_HBPmpE8Tzmzn0",   # 发送给董庆豪
        "template_id": "ksNf6WiqO5JEqd3bY6SUqJvWeL2-kEDqukQC4VeYVvw",
        "url": "http://wenda.zhugeyingxiao.com/api/tongji_kehu_shiyong",
        "data": {
            "first": {
                "value": "最近3天未操作的客户！",
                "color": "#173177"
            },
            "keyword1": {
                "value": datetime.datetime.now().strftime("%y-%m-%d %H:%M:%S"),
                "color": "#173177"
            },
        }
    }

    send_msg_obj = WeChat.WeChatPublicSendMsg()

    touser_list = [
        "o7Xw_0bdwjqmqSsXBVGfZiYMy0pQ",
        "o7Xw_0ZgxSLdvmtFdOd4c9W4OAn4",
        "o7Xw_0RM0Cw5IU_HBPmpE8Tzmzn0",
    ]
    for touser in touser_list:
        post_data["touser"] = touser
        send_msg_obj.sendTempMsg(post_data)


# 覆盖报表功能中生成客户覆盖报表
@app.task
def cover_reports_generate_excel(file_name, data_list, debug, url_list=None):
    # 生成客户查看的覆盖报表
    print('生成客户查看的覆盖报表 -->')
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="客户名称")
    ws.cell(row=1, column=2, value="关键词")
    ws.cell(row=1, column=3, value="覆盖类型")
    ws.cell(row=1, column=4, value="链接")
    ws.cell(row=1, column=5, value="排名")
    ws.cell(row=1, column=6, value="创建时间")
    if debug == False:
        ws.cell(row=1, column=7, value="类型")
        ws.cell(row=1, column=8, value="发布时间")
        ws.cell(row=1, column=9, value="问答类型")
    if url_list:
        ws.cell(row=1, column=10, value="链接占比")
        ws.cell(row=1, column=11, value="覆盖占比")
    if url_list:  # 如果有url_list 则表示需要填写新问答和老问答所占的百分比
        xinfugai_count, laofugai_count = 0, 0
        xinlianjie = len(set(url_list['xinlianjie']))
        laolianjie = len(set(url_list['laolianjie']))
        xinfugais = url_list['xinfugai']
        laofugais = url_list['laofugai']
        for xinfugai in xinfugais:
            xinfugai_count += xinfugai
        for laofugai in laofugais:
            laofugai_count += laofugai
        count_fugailiang = xinfugai_count + laofugai_count
        count_baifenbi = xinlianjie + laolianjie

        xinbaifenbi = float((xinlianjie + .0) / count_baifenbi) * 100
        laobaifenbi = float((laolianjie + .0) / count_baifenbi) * 100
        xinfugai = float((xinfugai_count + .0) / count_fugailiang) * 100
        laofugai = float((laofugai_count + .0) / count_fugailiang) * 100
        yunsuanxin = round(xinbaifenbi)
        yunsuanlao = round(laobaifenbi)
        xinfugaiyunsuan = round(xinfugai)
        laofugaiyunsuan = round(laofugai)
        ws.cell(row=2, column=10, value="新问答占比:" + str(yunsuanxin) + '%')
        ws.cell(row=3, column=10, value="老问答占比:" + str(yunsuanlao) + '%')
        ws.cell(row=2, column=11, value="新覆盖:" + str(xinfugaiyunsuan) + '%')
        ws.cell(row=3, column=11, value="老覆盖:" + str(laofugaiyunsuan) + '%')

    for row, i in enumerate(data_list, start=2):
        try:
            ws.cell(row=row, column=1, value=i["username"])
            ws.cell(row=row, column=2, value=i["keywords"])
            ws.cell(row=row, column=3, value=i["page_type"])
            if debug == False:
                ws.cell(row=row, column=4, value=i["link"])
                ws.cell(row=row, column=7, value=i["is_zhedie"])
                ws.cell(row=row, column=8, value=i["create_time"])
                ws.cell(row=row, column=9, value=i["wenda_type"])


            else:
                if i["link"]:
                    ws["D{row}".format(row=row)].hyperlink = i["link"]
                    ws["D{row}".format(row=row)].value = "点击打开知道问答页面"

            ws.cell(row=row, column=5, value=i["rank"])
            ws.cell(row=row, column=6, value=i["create_date"])

            row += 1
        except IllegalCharacterError:
            print("error -->", i)

    wb.save(file_name)


# 生成客户日覆盖报表   (老问答覆盖模式)
@app.task
def userprofile_keywords_cover(debug=False):
    client_data = models.KeywordsCover.objects.values(
        'keywords__client_user__username',
        'keywords__client_user_id'
    ).annotate(Count("id"))

    for user_obj in client_data:
        user_id = user_obj["keywords__client_user_id"]
        username = user_obj["keywords__client_user__username"]
        date_obj = datetime.datetime.now()
        date = date_obj.strftime("%Y-%m-%d")

        q = Q(Q(update_select_cover_date__isnull=True) | Q(update_select_cover_date__lt=date))
        keywords_topset_obj = models.KeywordsTopSet.objects.filter(
            client_user_id=user_id,
            is_delete=False
        ).filter(q)

        keyword_select_count = keywords_topset_obj.count()  # 未查询的关键词数

        # 十分钟之前有新数据则不生成报表
        ten_minutes_ago = date_obj - datetime.timedelta(minutes=10)
        keywords_cover_objs = models.KeywordsCover.objects.filter(
            keywords__client_user_id=user_id,
            create_date__gt=ten_minutes_ago
        )

        if keyword_select_count > 0 or keywords_cover_objs.count() > 0:  # 如果大于0表示数据未查询完
            continue

        user_profile_keywords_cover_obj = models.UserprofileKeywordsCover.objects.filter(
            client_user_id=user_id,
            create_date=date,
        )
        if user_profile_keywords_cover_obj:  # 如果已经创建,则跳过
            continue

        search_objs = models.KeywordsCover.objects.select_related(
            'keywords',
            'keywords__client_user'
        ).filter(
            keywords__client_user_id=user_id,
            create_date__year=date_obj.year,
            create_date__month=date_obj.month,
            create_date__day=date_obj.day,
        ).order_by("-create_date")

        if not search_objs:
            continue

        url_list = {
            'xinlianjie': [],
            'laolianjie': [],
            'xinfugai': [],
            'laofugai': []
        }

        data_day_list = [] # 新问答
        data_list1 = []  # 老问答不折叠
        data_list2 = []  # 其他
        temp_list = []
        for search_obj in search_objs:
            if search_obj:
                fugai_count = search_obj.keywords.top_page_cover
                print('search_obj - - - > ', search_obj.id)
                url = search_obj.url
                print('url_', url)
                print('search _ obj ', search_obj.keywords.client_user)
                objs = models.WendaRobotTask.objects.filter(
                    wenda_url=url,
                    task__release_user=search_obj.keywords.client_user
                )
                create_time = ''
                print('objs - -- - =--==- > ', objs)
                if objs:
                    if objs[0].wenda_type in [1, 10]:
                        url_list['xinlianjie'].append(url)
                        url_list['xinfugai'].append(fugai_count)
                    if objs[0].wenda_type == 2:
                        url_list['laolianjie'].append(url)
                        url_list['laofugai'].append(fugai_count)
                    robotaccountlog_objs = objs[0].robotaccountlog_set.all()
                    if robotaccountlog_objs:
                        create_time = robotaccountlog_objs.last().create_date.strftime("%Y-%m-%d")
                    is_zhedie = "0"
                    if search_obj.is_zhedie:
                        is_zhedie = "1"
                    wenda_type_index, wenda_type = objs[0].wenda_type, objs[0].get_wenda_type_display()
                    # data_list.append([is_zhedie,search_obj.url])
                    # data_day_list.append({
                    #     "username": username,
                    #     "keywords": search_obj.keywords.keyword,
                    #     "page_type": search_obj.get_page_type_display(),
                    #     "rank": search_obj.rank,
                    #     "create_date": search_obj.create_date.strftime("%Y-%m-%d"),
                    #     "link": search_obj.url,
                    #     "is_zhedie": is_zhedie,
                    #     'create_time': create_time,
                    #     'wenda_type': wenda_type
                    # })

                    line_data = {
                        "username": username,
                        "keywords": search_obj.keywords.keyword,
                        "page_type": search_obj.get_page_type_display(),
                        "rank": search_obj.rank,
                        "create_date": search_obj.create_date.strftime("%Y-%m-%d"),
                        "link": search_obj.url,
                        "is_zhedie": is_zhedie,
                        'create_time': create_time,
                        'wenda_type': wenda_type
                    }

                    # 新问答数据
                    if wenda_type_index in [1, 10]:
                        data_day_list.append(line_data)

                    # 老问答数据
                    else:
                        # 老问答未折叠
                        if is_zhedie == '0' and [search_obj.url, wenda_type] not in temp_list:
                            data_list1.append(line_data)
                            temp_list.append([search_obj.url, wenda_type])
                        else:
                            data_list2.append(line_data)

                    data_day_list.extend(data_list1)
                    data_day_list.extend(data_list2)

        # 客户查看报表的名称
        file_name = "{username}_{date}.xlsx".format(
            username=username,
            date=date
        )
        file_path_name = os.path.join("statics", "upload_files", file_name)

        cover_reports_generate_excel(file_path_name, data_day_list, debug=True)

        # 营销顾问查看报表的名称
        yingxiaoguwen_file_name = "yingxiaoguwen_{username}_{date}.xlsx".format(
            username=username,
            date=date
        )
        yingxiaoguwen_file_path_name = os.path.join("statics", "upload_files", yingxiaoguwen_file_name)

        cover_reports_generate_excel(yingxiaoguwen_file_path_name, data_day_list, debug=False, url_list=url_list)

        url_num = search_objs.values('url').distinct().count()

        print("生成报表 -->>")
        models.UserprofileKeywordsCover.objects.create(
            client_user_id=user_id,
            create_date=date,
            cover_num=len(data_day_list),
            statement_path=file_path_name,
            url_num=url_num
        )


# 每天提醒客户报表已经查完
@app.task
def send_cover_info():
    # 生成报表功能
    # 判断是否为调试模式 如果为调试模式不推送微信通知
    objs = models.GlobalSettings.objects.all()
    if objs[0].fugaibaobiao_shengcheng_moshi:
        return
    client_data = models.KeywordsCover.objects.filter(
        keywords__client_user__openid__isnull=False,  # 关注公众号
        keywords__client_user__send_statement=True,  # 发送报表
    ).values(
        'keywords__client_user_id', 'keywords__client_user__openid'
    ).annotate(Count("id"))

    now_date = datetime.datetime.now().strftime("%Y-%m-%d")

    we_chat_public_send_msg_obj = WeChat.WeChatPublicSendMsg()

    for obj in client_data:
        user_id = obj["keywords__client_user_id"]
        openid = obj["keywords__client_user__openid"]

        userprofile_keywords_cover_obj = models.UserprofileKeywordsCover.objects.filter(
            create_date=now_date,
            client_user_id=user_id,
            is_send_wechat=False,
            client_user__status=1  # 如果用户未启用，则不发送报表
        )

        # 如果不存在,则表示还未查询完
        if not userprofile_keywords_cover_obj:
            continue

        post_data = {
            "touser": openid,
            "template_id": "ksNf6WiqO5JEqd3bY6SUqJvWeL2-kEDqukQC4VeYVvw",
            "url": "http://wenda.zhugeyingxiao.com/show_wenda_cover_num/{openid}/{date_time}".format(
                openid=openid,
                date_time=datetime.datetime.now().strftime("%Y-%m-%d")
            ),
            "data": {
                "first": {
                    "value": "今日报表已生成",
                    "color": "#173177"
                },
                "keyword1": {
                    "value": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    # "color": "#173177"
                },
                "keyword2": {
                    "value": "诸葛问答",
                },
                # "keyword3": {
                #     "value": "发布失败",
                #     "color": "#173177"
                # },
                # "keyword4": {
                #     "value": "请修改",
                #     "color": "#173177"
                # },
                "remark": {
                    "value": "点击下方详情查看今日详细覆盖报表",
                    "color": "#173177"
                }
            }
        }
        we_chat_public_send_msg_obj.sendTempMsg(post_data)

        post_data = {
            "touser": 'o7Xw_0fq6LrmCjBbxAzDZHTbtQ3g',
            "template_id": "ksNf6WiqO5JEqd3bY6SUqJvWeL2-kEDqukQC4VeYVvw",
            "url": "http://wenda.zhugeyingxiao.com/show_wenda_cover_num/{openid}/{date_time}".format(
                openid=openid,
                date_time=datetime.datetime.now().strftime("%Y-%m-%d")
            ),
            "data": {
                "first": {
                    "value": "今日报表已生成",
                    "color": "#173177"
                },
                "keyword1": {
                    "value": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    # "color": "#173177"
                },
                "keyword2": {
                    "value": "诸葛问答",
                },
                # "keyword3": {
                #     "value": "发布失败",
                #     "color": "#173177"
                # },
                # "keyword4": {
                #     "value": "请修改",
                #     "color": "#173177"
                # },
                "remark": {
                    "value": "点击下方详情查看今日详细覆盖报表",
                    "color": "#173177"
                }
            }
        }
        we_chat_public_send_msg_obj.sendTempMsg(post_data)


        userprofile_keywords_cover_obj.update(is_send_wechat=True)


# 缓存指定首页关键词中的数据
@app.task
def cache_keywords_top_set_init_data():
    from webadmin.views_dir.wenda.keywords_top_set import init_data
    result_data = init_data()
    RedisOper.write_to_cache("keywords_top_set-init-data", result_data)


# 定时更新覆盖报表中的数据
@app.task
def update_client_covering_data():
    data_objs = models.KeywordsCover.objects.select_related(
        "keywords__client_user",
        "keywords",
    ).values(
        "keywords__client_user__username",
        "keywords__client_user_id",
    ).annotate(Count('id'))

    print(data_objs)

    for obj in data_objs:
        client_user_id = obj['keywords__client_user_id']

        # 总覆盖数
        total_cover_num = 0
        if client_user_id == 140:  # 西宁东方泌尿专科 删除词了,之前对应的覆盖也删除了,单独处理
            userprofile_keywords_cover_objs = models.UserprofileKeywordsCover.objects.filter(
                client_user_id=client_user_id)
            for userprofile_keywords_cover_obj in userprofile_keywords_cover_objs:
                total_cover_num += userprofile_keywords_cover_obj.cover_num

        else:
            total_cover_num = obj['id__count']

        keywords_topset_obj = models.KeywordsTopSet.objects.filter(
            client_user_id=obj["keywords__client_user_id"],
            is_delete=False
        )

        # 关键词总数
        keyword_count = keywords_topset_obj.count()

        now_date = datetime.datetime.now().strftime("%Y-%m-%d")
        # 今日覆盖数
        today_cover_num = models.KeywordsCover.objects.filter(
            keywords__client_user_id=client_user_id,
            create_date__gte=now_date
        ).count()

        # 总发布次数
        total_publish_num = models.RobotAccountLog.objects.filter(
            wenda_robot_task__task__release_user_id=client_user_id,
            wenda_robot_task__wenda_type=2
        ).count()

        q = Q(Q(update_select_cover_date__isnull=True) | Q(update_select_cover_date__lt=now_date))

        # 未查询的关键词数
        keyword_no_select_count = keywords_topset_obj.filter(q).count()

        data = {
            "keywords_num": keyword_count,
            "keyword_no_select_count": keyword_no_select_count,
            "today_cover_num": today_cover_num,
            "total_cover_num": total_cover_num,
            "total_publish_num": total_publish_num,
            "update_date": datetime.datetime.now()
        }

        client_covering_data_objs = models.ClientCoveringData.objects.filter(client_user_id=client_user_id)
        if client_covering_data_objs:
            client_covering_data_objs.update(**data)
        else:
            data['client_user_id'] = client_user_id

            models.ClientCoveringData.objects.create(**data)


# 从问答采集后台获取数据插入到养账号任务中
@app.task
def yangzhanghao_task():
    api_url = "http://wenda.questions.zhugeyingxiao.com/api/questions"

    ret = requests.get(api_url)
    result_data = ret.json()
    import datetime

    if result_data['status'] and result_data['data']:
        query = []
        for data in result_data['data']:
            tid = data['tid']
            title = data['title']
            content = data['content']
            query.append(
                models.WendaRobotTask(
                    task_id=734,
                    title=title,
                    content=content,
                    release_platform=1,
                    wenda_type=1,
                    next_date=datetime.datetime.now()
                )
            )
        models.WendaRobotTask.objects.bulk_create(query)


# 将最近6个小时使用的 ip 地址缓存到redis 中
@app.task
def cached_ipaddr_list():
    from webadmin.modules import RedisOper
    up_hours_time = datetime.datetime.now() - datetime.timedelta(hours=6)
    ipaddr_list = [i[0] for i in
                   models.RobotAccountLog.objects.filter(create_date__lt=up_hours_time).values_list('ipaddr')]
    RedisOper.write_to_cache("api_check_ipaddr_ip_list", list(ipaddr_list))


# 更新问答库编辑编写数据
@app.task
def update_wendaku_bianjibianxie():
    token = '4e0398e4d4bad913e24c1d0d60cc9170'
    timestamp = str(int(time.time()))

    # 用户输入的密码加密
    pwd = str(timestamp + token)
    hash = hashlib.md5()
    hash.update(pwd.encode())
    str_encrypt = hash.hexdigest()
    params = {
        'user_id': 1,
        'timestamp': timestamp,
        'rand_str': str_encrypt
    }

    params['is_today'] = True

    # url = 'http://192.168.10.243:8000/wendaku/cishu'
    url = 'http://api.zhugeyingxiao.com/wendaku/cishu'
    # ret = requests.post('http://192.168.10.243:8000/ribao/user/add/0', params=params, data=data)
    ret = requests.get(url, params=params)
    data_temp = ret.json()
    json_data = data_temp['data']
    data_list = []
    for k1, v1 in json_data.items():
        # print('v--->',v1)
        for k2, v2 in v1.items():
            # print('v1--v2-->',k2,v2)
            objs = models.UserProfile.objects.filter(username=v2[0], is_delete=False)
            if objs:
                obj = objs[0]
                # print('obj_id-->',obj.id)
                p = v2[1]
                print('k2-->', k2)

                bianxiebaobiao_objs = models.BianXieBaoBiao.objects.filter(
                    xiangmu=1,
                    oper_user_id=obj.id,
                    create_date=k2,
                )
                if bianxiebaobiao_objs:
                    print('修改数据---->', bianxiebaobiao_objs)
                    bianxiebaobiao_objs.update(
                        edit_count=p,
                    )

                else:
                    models.BianXieBaoBiao.objects.create(
                        xiangmu=1,
                        oper_user_id=obj.id,
                        create_date=k2,
                        edit_count=p,
                    )

            else:
                print(v2[0], '不存在')


# 更新编辑编写新老问答数据
@app.task
def update_xinlaowenda_bianxie_cishu():
    now_date = datetime.datetime.now().strftime('%Y-%m-%d')
    objs = models.EditPublickTaskManagement.objects.filter(create_date__gte=now_date).values('task__edit_user',
        'task__edit_user__username', 'task__task__task__wenda_type', 'create_date').annotate(Count('id'))

    data_dict = {
        'xin': {},
        'lao': {},
    }

    for obj in objs:
        # print('obj --> ',obj)
        user_id = obj['task__edit_user']
        username = obj['task__edit_user__username']
        id_count = obj['id__count']
        create_date = obj['create_date'].strftime('%Y-%m-%d')
        task__wenda_type = obj['task__task__task__wenda_type']
        # print('数据--->',username,user_id,task__wenda_type,create_date,id_count)

        if task__wenda_type == 2:  # 老问答
            if user_id in data_dict['lao']:
                if create_date in data_dict['lao'][user_id]:
                    data_dict['lao'][user_id][create_date] += id_count
                else:
                    data_dict['lao'][user_id][create_date] = id_count
            else:
                data_dict['lao'][user_id] = {
                    # 'name': username,
                    create_date: id_count
                }

        else:  # 新问答
            if user_id in data_dict['xin']:
                if create_date in data_dict['xin'][user_id]:
                    data_dict['xin'][user_id][create_date] += id_count
                else:
                    data_dict['xin'][user_id][create_date] = id_count
            else:
                data_dict['xin'][user_id] = {
                    # 'name': username,
                    create_date: id_count
                }

    print('data_dict-->', data_dict)

    for k, v in data_dict.items():
        # print(k, v)
        if k == 'lao':
            xiangmu = 4
            for k2, v2 in v.items():
                # print(k2, v2)
                user_id = k2
                for k3, v3 in v2.items():
                    print(k3, v3)
                    create_date = k3
                    edit_count = v3

                    bianxiebaobiao_objs = models.BianXieBaoBiao.objects.filter(
                        xiangmu=xiangmu,
                        oper_user_id=user_id,
                        create_date=create_date,
                    )
                    if bianxiebaobiao_objs:
                        print('修改数据---->', bianxiebaobiao_objs)
                        bianxiebaobiao_objs.update(
                            edit_count=edit_count,
                        )

                    else:
                        models.BianXieBaoBiao.objects.create(
                            xiangmu=xiangmu,
                            oper_user_id=user_id,
                            create_date=create_date,
                            edit_count=edit_count,
                        )

        else:
            xiangmu = 3
            for k2, v2 in v.items():
                # print(k2, v2)
                user_id = k2
                for k3, v3 in v2.items():
                    print(k3, v3)
                    create_date = k3
                    edit_count = v3

                    print(xiangmu, user_id, create_date)
                    bianxiebaobiao_objs = models.BianXieBaoBiao.objects.filter(
                        xiangmu=xiangmu,
                        oper_user_id=user_id,
                        create_date=create_date,
                    )
                    if bianxiebaobiao_objs:
                        print('修改数据---->', bianxiebaobiao_objs)
                        bianxiebaobiao_objs.update(
                            edit_count=edit_count,
                        )

                    else:
                        models.BianXieBaoBiao.objects.create(
                            xiangmu=xiangmu,
                            oper_user_id=user_id,
                            create_date=create_date,
                            edit_count=edit_count,
                        )


# 修改编辑编写养账号数据
@app.task
def update_bianji_yangzhanghao_data():
    # q = Q()
    # q.add(Q(**{'status':2}),Q.AND)
    now_date = datetime.datetime.now().strftime('%Y-%m-%d')
    objs = models.ZhidaoWenda.objects.filter(status=2, update_date__gte=now_date).select_related('oper_user').values(
        'oper_user_id',
        'oper_user__username',
        'update_date',
    ).annotate(Count('id'))
    # print('objs--->',objs)
    data_dict = {}
    for obj in objs:
        count = obj['id__count']
        user_id = obj['oper_user_id']
        username = obj['oper_user__username']
        update_time = obj['update_date'].strftime('%Y-%m-%d')
        print('数据-->', count, user_id, username, update_time)

        if not user_id:
            continue

        if user_id in data_dict:
            if update_time in data_dict[user_id]:
                data_dict[user_id][update_time] += count

            else:
                data_dict[user_id][update_time] = count
        else:
            data_dict[user_id] = {
                # 'name':username,
                update_time: count
            }

    print('data_dict--->', data_dict)

    xiangmu = 2
    for k2, v2 in data_dict.items():
        # print(k2, v2)
        user_id = k2
        for k3, v3 in v2.items():
            print(k3, v3)
            create_date = k3
            edit_count = v3

            print(xiangmu, user_id, create_date)
            bianxiebaobiao_objs = models.BianXieBaoBiao.objects.filter(
                xiangmu=xiangmu,
                oper_user_id=user_id,
                create_date=create_date,
            )
            if bianxiebaobiao_objs:
                print('修改数据---->', bianxiebaobiao_objs)
                bianxiebaobiao_objs.update(
                    edit_count=edit_count,
                )

            else:
                models.BianXieBaoBiao.objects.create(
                    xiangmu=xiangmu,
                    oper_user_id=user_id,
                    create_date=create_date,
                    edit_count=edit_count,
                )


# 更新编辑更新新老问答数据
@app.task
def update_EditTaskLog_dahui_cishu():
    now_date = datetime.datetime.today().strftime('%Y-%m-%d')
    # print('now_date---->',now_date)
    objs = models.EditTaskLog.objects.filter(create_date__gte=now_date).values(
        'create_date',
        'edit_public_task_management__task__edit_user',
        'edit_public_task_management__task__task__task__wenda_type',
        'bianji_dahui_update'
    ).annotate(Count('id'))

    data_dict = {
        'xin_dahui': {},
        'lao_dahui': {},
        'xin_xiugai': {},
        'lao_xiugai': {},
    }

    for obj in objs:
        print('obj -- > ', obj)
        create_time = obj['create_date'].strftime('%Y-%m-%d')
        edit_user_id = obj['edit_public_task_management__task__edit_user']
        wenda_type = obj['edit_public_task_management__task__task__task__wenda_type']
        count = obj['id__count']
        bianji_dahui_update = obj['bianji_dahui_update']
        print('数据---->', create_time, edit_user_id, wenda_type, count)

        type_oper = ""
        if bianji_dahui_update == 1:
            type_oper = "dahui"
        elif bianji_dahui_update == 2:
            type_oper = "xiugai"

        if wenda_type == 2:  # 老问答
            # print('wenda_type ------- > ',wenda_type)
            type_oper = 'lao_' + type_oper

            if edit_user_id in data_dict[type_oper]:
                if create_time in data_dict[type_oper][edit_user_id]:
                    data_dict[type_oper][edit_user_id][create_time] += count
                else:
                    data_dict[type_oper][edit_user_id][create_time] = count

            else:
                data_dict[type_oper][edit_user_id] = {
                    create_time: count,
                }
        else:  # 新问答
            type_oper = 'xin_' + type_oper
            if edit_user_id in data_dict[type_oper]:
                if create_time in data_dict[type_oper][edit_user_id]:
                    data_dict[type_oper][edit_user_id][create_time] += count
                else:
                    data_dict[type_oper][edit_user_id][create_time] = count
            else:
                data_dict[type_oper][edit_user_id] = {
                    create_time: count,
                }

    for k1, v1 in data_dict.items():
        print('k1,v1  --- >', k1, v1)
        xiangmu = 0
        if k1 == 'xin_dahui':
            xiangmu = 5
        elif k1 == 'lao_dahui':
            xiangmu = 6
        elif k1 == 'xin_xiugai':
            xiangmu = 7
        elif k1 == 'lao_xiugai':
            xiangmu = 8

        for k2, v2 in v1.items():
            print('k2,v2- - - ->', k2, v2)
            for k3, v3 in v2.items():
                # print('k3,v3 --> ',k3,v3)

                obj = models.BianXieBaoBiao.objects.filter(
                    xiangmu=xiangmu,
                    create_date=k3,
                    oper_user_id=k2,
                )
                if obj:
                    # 如果存在只修改 总数
                    obj.update(edit_count=v3)
                else:
                    obj.create(
                        xiangmu=xiangmu,
                        create_date=k3,
                        edit_count=v3,
                        oper_user_id=k2,
                    )


# 微信推送到期用户
@app.task
def weixin_daoqi_yonghu_tuisong():
    webchat_obj = WeChatPublicSendMsg()
    # url = 'http://127.0.0.1:8006/api/jifeidaoqitixing/null/0'
    url = 'http://wenda.zhugeyingxiao.com/api/jifeidaoqitixing/null/0?canshu=1'
    ret = requests.get(url)
    ret_json = ret.content.decode()
    json_data = json.loads(ret_json)
    if json_data['data']:
        print('进入============')
        data_params = json_data['data']
        data_list = []
        for params in data_params:
            # 销售 发送
            # print('params= =======> ',params)
            if 'xiaoshou_openid' in params:
                data_list.append({
                    # xiaoshou_openid = params['xiaoshou_openid'],
                    'openid': 'o7Xw_0fq6LrmCjBbxAzDZHTbtQ3g',
                    'this_id': params['this_id'],
                    'stop_time': params['stop_time'],
                    'daoqi_today': params['daoqi_today'],
                })
            # 顾问
            else:
                data_list.append({
                    # guwen_openid = params['guwen_openid'],
                    'openid': 'o7Xw_0fq6LrmCjBbxAzDZHTbtQ3g',
                    'this_id': params['this_id'],
                    'daoqi_today': params['daoqi_today'],
                    'stop_time': params['stop_time']
                })


        for data in data_list:
            print('data ============ >', data)
            post_data = {
                "touser": "o7Xw_0fq6LrmCjBbxAzDZHTbtQ3g.",
                # "touser": "{openid}".format(openid=openid),
                "template_id": "ksNf6WiqO5JEqd3bY6SUqJvWeL2-kEDqukQC4VeYVvw",
                # "url": "http://wenda.zhugeyingxiao.com/api/jifeidaoqitixing/null/{gongyong_id}".format(gongyong_id=gongyong_id),
                "url": "http://127.0.0.1:8006/api/jifeidaoqitixing/null/{gongyong_id}".format(
                    gongyong_id=data['this_id']),
                "data": {
                    "first": {
                        "value": "诸葛霸屏王提醒有计费到期！",
                        "color": "#000"
                    },
                    "keyword1": {
                        "value": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        # "color": "#173177"
                    },
                    "keyword2": {
                        "value": "请见详情页面",
                    },
                }
            }
            print('post_data', post_data['url'])
            print('---------========================')
            webchat_obj.sendTempMsg(post_data)



# 微信推送每日覆盖量
@app.task
def weixin_meiri_fugai_tuisong():
    # 调用发送微信公众号模块
    webchat_obj = WeChatPublicSendMsg()
    now_date = datetime.datetime.today().strftime('%Y-%m-%d')
    q = Q()

    # 公用判断覆盖是否有值
    def gongyong_panduan(o_id):
        # print('公用判断')
        now_date = datetime.datetime.today().strftime('%Y-%m-%d')
        q = Q()
        q.add(Q(client_user__is_delete=False) & Q(client_user__status=1) & Q(create_date__gte=now_date), Q.AND)
        q.add(Q(client_user__xiaoshou__isnull=False) | Q(client_user__guwen__isnull=False), Q.AND)
        q.add(Q(client_user_id=o_id), Q.AND)
        objs = models.UserprofileKeywordsCover.objects.select_related('client_user').filter(q).values(
            'create_date',
            'cover_num',
            'client_user__username',
        ).annotate(Count('id'))
        print('objs - - -> ', objs)
        data_list = []
        for obj in objs:
            client_name = obj['client_user__username']
            cover_num = obj['cover_num']
            data_list.append({
                'name': client_name,
                'count': cover_num
            })
        print('data_list - -- -- > ', data_list)
        return data_list

    # 公用发送链接
    def gongyong(openid, gongyong_id, username):
        post_data = {
            "touser": "o7Xw_0fq6LrmCjBbxAzDZHTbtQ3g",
            # "touser": "{openid}".format(openid=openid),
            "template_id": "ksNf6WiqO5JEqd3bY6SUqJvWeL2-kEDqukQC4VeYVvw",
            "url": "http://wenda.zhugeyingxiao.com/api/fugailiangtixing/null/{gongyong_id}".format(
                gongyong_id=gongyong_id),
            # "url": "http://127.0.0.1:8006/api/fugailiangtixing/null/{gongyong_id}".format(gongyong_id=gongyong_id),
            "data": {
                "first": {
                    "value": "诸葛霸屏王查询覆盖通知！",
                    "color": "#000"
                },
                "keyword1": {
                    "value": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    # "color": "#173177"
                },
                "keyword2": {
                    "value": "覆盖详情请点击！",
                },
                "remark": {
                    "value": "查询用户:{username}".format(username=username),
                    "color": "#173177"
                }

            }
        }
        print('post_data', post_data['url'])
        print('---------========================')
        webchat_obj.sendTempMsg(post_data)

    # 顾问
    def guwen_weixin():
        q.add(Q(client_user__is_delete=False) & Q(client_user__status=1) & Q(create_date__gte=now_date) & (
            Q(client_user__guwen__isnull=False)), Q.AND)
        # q.add(Q(client_user__is_delete=False) & Q(client_user__status=1) & (Q(client_user__guwen__isnull=False)), Q.AND)
        objs = models.UserprofileKeywordsCover.objects.filter(q).order_by('create_date')
        if objs:
            yonghu_id = objs[0].client_user.id
            data_list = gongyong_panduan(yonghu_id)
            data_temp = {}
            for obj in objs:
                guwen_openid = obj.client_user.guwen.openid
                username = obj.client_user.username
                data_temp['openid'] = guwen_openid
                data_temp['username'] = username
            if data_list:
                gongyong(data_temp['openid'], yonghu_id, data_temp['username'])

    # 销售
    def xiaoshou_weixin():
        q.add(Q(client_user__is_delete=False) & Q(client_user__status=1) & Q(create_date=now_date) & (
            Q(client_user__xiaoshou__isnull=False)), Q.AND)
        objs = models.UserprofileKeywordsCover.objects.select_related('client_user').filter(q).order_by('create_date')
        if objs:
            yonghu_id = objs[0].client_user.id
            data_list = gongyong_panduan(yonghu_id)
            data_temp = {}
            for obj in objs:
                xiaoshou_openid = obj.client_user.xiaoshou.openid
                username = obj.client_user.username
                data_temp['openid'] = xiaoshou_openid
                data_temp['username'] = username
            if data_list:
                gongyong(data_temp['openid'], yonghu_id, data_temp['username'])

    # print('顾问 - -- - 》 ')
    # guwen_weixin()
    print('销售 - - -- 》 ')
    xiaoshou_weixin()


# 更新机器人日志发布次数
def robot_release_num():
    now_date = datetime.datetime.today().strftime('%Y-%m-%d %H:00:00')
    objs = models.RobotAccountLog.objects.filter(
        lapse=False,
        # create_date__gte = now_date
    ).values(
        'status',
        'create_date',
    ).annotate(Count('id'))
    date_temp = {}
    two_temp = {}
    for obj in objs:
        status = obj['status']
        create_time = obj['create_date'].strftime('%Y-%m-%d %H')
        id_count = obj['id__count']
        if status in date_temp:
            if create_time in date_temp[status]:
                date_temp[status][create_time] += id_count
            else:
                date_temp[status][create_time] = id_count
        else:
            date_temp[status] = {
                create_time: id_count
            }
        for key, val in date_temp.items():
            for key1, val1 in val.items():
                data_key = key1 + ':00:00'
                if data_key in two_temp:
                    if val1 in two_temp[data_key]:
                        two_temp[data_key]['count'] += val1
                    else:
                        two_temp[data_key]['count'] = val1
                else:
                    two_temp[data_key] = {
                        'count': val1
                    }
    for k, v in two_temp.items():
        for k1, v1 in v.items():
            create_date = k
            count = v1
            objs = models.RobotReleaseNum.objects.filter(
                create_date=create_date,
            )
            if objs:
                objs.update(
                    create_date=create_date,
                    robot_count=count
                )
            else:
                models.RobotReleaseNum.objects.create(
                    create_date=create_date,
                    robot_count=count
                )


# 下载关键词与关键词类型
@app.task
def guanjianci_xiazai(file_name, data_list):
    wb = Workbook()
    ws = wb.active
    print('进入下载 关键词 ')
    ws.cell(row=1, column=1, value="关键词")
    ws.cell(row=1, column=2, value="关键词类型")
    for row, i in enumerate(data_list, start=2):
        try:
            ws.cell(row=row, column=1, value=i["keyword"])
            ws.cell(row=row, column=2, value=i["keywords_type"])
            row += 1
        except IllegalCharacterError:
            print("error -->", i)
    wb.save(file_name)


# 宕机微信推送提醒
@app.task
def dangjitixing():
    webchat_obj = WeChatPublicSendMsg()
    # 请求得URL获取备注宕机信息
    url = 'http://websiteaccount.bjhzkq.com/api/checkVpsStatus'
    # 请求URL的参数
    ret = requests.get(url)
    if ret.text:
        post_data = {
            # 李汉杰openid
            "touser": "o7Xw_0c264_Xjns8vKxHaFakAfIw",
            # 自己openid
            # "touser": "o7Xw_0fq6LrmCjBbxAzDZHTbtQ3g",
            "template_id": "JHyJHGz4QP_wpe67yucklAl3KM1tdWxbnmcqNLPhZaU",
            "data": {
                "first": {
                    "value": "宕机提醒,请及时处理！",
                    "color": "#000"
                },
                "time": {
                    "value": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    # "color": "#173177"
                },
                "reason": {
                    "value": "{}".format(ret.text),
                },
            }
        }
        print('---------========================')
        webchat_obj.sendTempMsg(post_data)


# 新问答完成的不打回到编辑
@app.task
def xinwenda_wancheng_budahui():
    requests.get('http://wenda.zhugeyingxiao.com/api/xinwenda_wancheng_budahui')


# 指定关键词-优化-协助调用查询数据库
@app.task
def keywords_select_models():
    # url = 'http://api.zhugeyingxiao.com/api/keywords_select_models'
    # # url = 'http://127.0.0.1:8006/api/keywords_select_models'
    # data = {
    #     'canshu': 'KeywordsTopInfo'
    # }
    # user_id_list, user_data = [], {}
    # ret = requests.post(url, data=data)
    # sleep(1)
    # data_objs = ret.text
    # json_data_objs = json.loads(data_objs)
    # data_objs = json_data_objs['data']['data_objs_list']
    #
    # for obj in data_objs:
    #     client_user_id = int(obj["keyword__client_user"])
    #     username = obj["keyword__client_user__username"]
    #     page_type = obj["page_type"]
    #     cover = obj["cover"]
    #
    #     if client_user_id in user_id_list:
    #         user_data[client_user_id][page_type] = cover
    #     else:
    #         user_id_list.append(client_user_id)
    #         user_data[client_user_id] = {
    #             page_type: cover,
    #             "username": username,
    #             "user_id": client_user_id
    #         }
    # for index, user_id in enumerate(user_id_list):
    #     client_user_id = user_data[user_id]["user_id"]
    #     username = user_data[user_id]["username"]
    #     pc_cover = 0
    #     wap_cover = 0
    #     if 1 in user_data[user_id]:
    #         pc_cover = user_data[user_id][1]
    #     if 3 in user_data[user_id]:
    #         wap_cover = user_data[user_id][3]
    #     total_cover = pc_cover + wap_cover
    #     data = {
    #         'canshu': 'KeywordsTopSet',
    #         'user_id': client_user_id
    #     }
    #     ret = requests.post(url, data=data)
    #     sleep(1)
    #     data_str = json.loads(ret.text)['data']
    #     keywords_num = data_str['keywords_num']
    #     no_select_keywords_num = data_str['no_select_keywords_num']
    #     keywords_top_page_cover_excel_path = data_str['keywords_top_page_cover_excel_path']
    #     keywords_top_page_cover_yingxiao_excel_path = data_str['keywords_top_page_cover_yingxiao_excel_path']
    #     if no_select_keywords_num > 0:
    #         keywords_status = '1'
    #     else:
    #         keywords_status = '2'
    #     data = {
    #         'canshu': 'UserProfile',
    #         'user_id': client_user_id
    #     }
    #     ret = requests.post(url, data=data)
    #     sleep(1)
    #     user_obj_id = json.loads(ret.text)['data']['user_obj']
    #     data_temp = {
    #         'username_id': user_obj_id,
    #         'koywords_status': keywords_status,
    #         'keywords_num': keywords_num,
    #         'total_cover': total_cover,
    #         'pc_cover': pc_cover,
    #         'wap_cover': wap_cover,
    #         'no_select_keywords_num': no_select_keywords_num,
    #         'keywords_top_page_cover_excel_path': keywords_top_page_cover_excel_path,
    #         'keywords_top_page_cover_yingxiao_excel_path': keywords_top_page_cover_yingxiao_excel_path
    #     }
    #     requests.post(url, data=data_temp)
    url= 'http://wenda.zhugeyingxiao.com/api/keywords_select_models'
    ret = requests.get(url)
    print(ret.text)








