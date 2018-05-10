#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author:zhangcong
# Email:zc_92@sina.com

import os
import sys

project_dir = os.path.dirname(os.getcwd())

sys.path.append(project_dir)
print(project_dir)

os.environ['DJANGO_SETTINGS_MODULE'] ='wenda.settings'
import django
django.setup()

from webadmin import models
from openpyxl import Workbook
import time

from openpyxl.utils.exceptions import IllegalCharacterError

def keywords_top_page_cover_excel():
    user_profile_objs = models.UserProfile.objects.filter(role_id=5)

    for user_obj in user_profile_objs:

        # 如果没有报表存在,则生成报表
        if not user_obj.keywords_top_page_cover_excel_path:
            keywords_top_set_objs = models.KeywordsTopSet.objects.filter(client_user=user_obj)

            # 没有指定首页关键词的跳过
            if keywords_top_set_objs.count() == 0:
                continue
            else:

                # 假如还有关键没有查询完
                if keywords_top_set_objs.filter(status=1).count() > 0:
                    continue

                else:
                    wb = Workbook()
                    ws = wb.active

                    ws.cell(row=1, column=1, value="编号")
                    ws.cell(row=1, column=2, value="客户名称")
                    ws.cell(row=1, column=3, value="搜索关键词")
                    ws.cell(row=1, column=4, value="知道标题")
                    ws.cell(row=1, column=5, value="知道链接")
                    ws.cell(row=1, column=6, value="排名位置")
                    ws.cell(row=1, column=7, value="排名类型")
                    ws.cell(row=1, column=8, value="是否采纳")
                    ws.cell(row=1, column=9, value="创建时间")

                    row = 2
                    for keywords_top_set_obj in keywords_top_set_objs:
                        for obj in keywords_top_set_obj.keywordstopinfo_set.all():
                            try:

                                if obj.is_caina:
                                    is_caina = "已采纳"
                                else:
                                    is_caina = "未采纳"

                                ws.cell(row=row, column=1, value=row - 1)
                                ws.cell(row=row, column=2, value=user_obj.username)
                                ws.cell(row=row, column=3, value=keywords_top_set_obj.keyword)
                                ws.cell(row=row, column=4, value=obj.title)
                                ws.cell(row=row, column=5, value=obj.url)
                                ws.cell(row=row, column=6, value=obj.rank)
                                ws.cell(row=row, column=7, value=obj.get_page_type_display())
                                ws.cell(row=row, column=8, value=is_caina)
                                ws.cell(row=row, column=9, value=keywords_top_set_obj.create_date.strftime("%y-%m-%d %H:%M:%S"))

                                row += 1
                            except IllegalCharacterError:
                                print(keywords_top_set_obj.keyword)

                    excel_path = "statics/task_excel/keywords_top_set/{name}.xlsx".format(name=str(int(time.time() * 1000)))
                    print(excel_path)
                    wb.save(os.path.abspath(excel_path))
                    # user_obj.keywords_top_page_cover_excel_path = excel_path
                    # user_obj.save()

if __name__ == '__main__':
    keywords_top_page_cover_excel()

